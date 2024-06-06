from  resources_positons import  doctors_list, nurses_list, patients_list,devices_list,names_cards
from indications_positions import temperatures,humidity
from indications_positions import details_pos
from rooms_design_positions import corridors,rooms
import time
from openpyxl import Workbook,load_workbook
import openpyxl
from datetime import datetime
import pygame
import sys
import mysql.connector


red = (255, 0, 0)
green = (0, 255, 0)

_id_for_delete = " "

#at each  room put the idσ cards and type of card
def readCards(string,room_card_list,connections_ok,connection_ok_for_print):
     print(string)
     room = string.split(':')[0]
     cards = string.split(':')[1]
     
     if room == "1" :
         connections_ok[0]=True
         connection_ok_for_print[0]=True
         print("room 1")
         writeExcel(room,room_card_list,cards)
         deleteDublicate(room,room_card_list,cards)
         room_card_list["room1"]={}
         if  cards:
             for key in cards.split('|'):
                 room_card_list["room1"][key]= findType(key)


         corridors[8]['alarm'] = string.split(':')[4] 
         if string.split(':')[4] == "1":
             corridors[8]['lockAlarm']= "1"
         rooms[0]['fire']= string.split(':')[5]
         corridors[14]['alarm']= string.split(':')[6]
          
         temperatures[0]['value'] =string.split(':')[2][:-1]
         humidity[0]['value'] =string.split(':')[3].split('.')[0]
         

     elif room =="2" :
         connections_ok[1]=True
         connection_ok_for_print[1]=True
         	
         print("room 2")
         writeExcel(room,room_card_list,cards)
         deleteDublicate(room,room_card_list,cards)
         room_card_list["room2"]={}
         if  cards:
              for key in cards.split('|'):
                  room_card_list["room2"][key]= findType(key)
         
         corridors[9]['alarm'] = string.split(':')[4] 
         if string.split(':')[4] == "1":
             corridors[9]['lockAlarm']= "1"
         rooms[1]['fire']= string.split(':')[5]
         corridors[15]['alarm'] = string.split(':')[6]  
       
         
         temperatures[1]['value'] =22.5
         humidity[1]['value'] =60


     elif room =="3" :
          connections_ok[2]=True
          connection_ok_for_print[2]=True

          print("room 3")
          writeExcel(room,room_card_list,cards)
          deleteDublicate(room,room_card_list,cards)
          room_card_list["room3"]={}
          if cards:
                for key in cards.split('|'):
                  room_card_list["room3"][key]= findType(key)
     
          corridors[10]['alarm'] = string.split(':')[4]
          if string.split(':')[4] == "1":
             corridors[10]['lockAlarm']= "1"
          rooms[2]['fire']= string.split(':')[5] 
          corridors[16]['alarm'] = string.split(':')[6] 
               
          temperatures[2]['value'] =22.0
          humidity[2]['value'] =62
     

     elif room =="4" :
         connections_ok[3]=True
         print("room 4")
         room_card_list["room4"]={}
         if  cards:
              for key in cards.split('|'):
                  room_card_list["room4"][key]= findType(key)
    
         corridors[11]['alarm'] = string.split(':')[4]
         rooms[3]['fire']= string.split(':')[5]
         corridors[17]['alarm'] = string.split(':')[6] 
           
         temperatures[3]['value'] =19.9
         humidity[3]['value'] =70


     elif room =="5" :
          connections_ok[4]=True
          print("room 5")
          room_card_list["room5"]={}
          if cards:
                for key in cards.split('|'):
                  room_card_list["room5"][key]= findType(key)
     
          corridors[12]['alarm'] = string.split(':')[4]
          rooms[4]['fire']= string.split(':')[5]
          corridors[18]['alarm'] = string.split(':')[6] 
          
          temperatures[4]['value'] =22.4
          humidity[4]['value'] =58


     elif room =="6" :   
          connections_ok[3]=True
          connection_ok_for_print[3]=True
          
          
          print("room 6")
          writeExcel(room,room_card_list,cards)
          print("cards : ",cards)
          deleteDublicate(room,room_card_list,cards)
          room_card_list["room6"]={}
          
          if cards:
                for key in cards.split('|'):
                  room_card_list["room6"][key]= findType(key)
          
          corridors[13]['alarm'] = string.split(':')[4]
          if string.split(':')[4]=="1":
               corridors[13]['lockAlarm']="1"
          rooms[5]['fire']= string.split(':')[5]
          corridors[19]['alarm'] = string.split(':')[6]
                  
          temperatures[5]['value'] =22.0
          humidity[5]['value'] =60
      
          temperatures[6]['value'] =21.3
          humidity[6]['value'] =62

     elif room =="sec" :
         connection_ok_for_print[1]=True
         if  string.split(':')[1] == "1":
             return green
         else:
             return red

     else :
        
        print("error")
        #connections_ok[6]=False

     
     print(" priexomena domayioy  1: ",room_card_list["room1"])
     print(" priexomena domayioy  2: ",room_card_list["room2"])
     print(" priexomena domayioy  3: ",room_card_list["room3"])
     print(" priexomena domayioy  4: ",room_card_list["room4"])
     print(" priexomena domayioy  5: ",room_card_list["room5"])
     print(" priexomena domayioy  6: ",room_card_list["room6"])
     

     
#find what type is each  card is 
def findType(key):
        
        if   key in doctors_list:
             return "doctor"
        elif key in nurses_list:
             return "nurse"
        elif key in patients_list:
             return "patient"
        elif key in devices_list:
             return "device"
        else:
             return "error"


def details( screen,room_card_list):
     
     
     GREW = (185, 204, 179)  # Μπλε χρώμα για το τετράγωνο
    
     DOC = (0, 122, 0) 
     NUR= (255, 0, 80)
     PAT= (0, 0, 255)
     DEV= (204, 102, 0)
     
     DOCC = (0, 122, 0) 
     NURR= (255, 0, 80)
     PATT= (0, 0, 255)
     DEVV= (204, 102, 0)

     room6_doctors = list({names_cards[key] for key, value in room_card_list["room6"].items() if value == "doctor"})
     room6_nurses = list({names_cards[key] for key, value in room_card_list["room6"].items() if value == "nurse"})
     room6_patients = list({names_cards[key] for key, value in room_card_list["room6"].items() if value == "patient"})
     room6_devices = list({names_cards[key] for key, value in room_card_list["room6"].items() if value == "device"})
     for (room,card_dict),details in  zip(room_card_list.items(),details_pos):
          #print(f"Επεξεργασία του δωματίου: {room}")
          if room != "room6":
               doctors = []
               nurses = []
               patients= [] 
               devices = []
               
               for key,value in card_dict.items():
                    if value == 'doctor':
                       doctors.append(names_cards[key])
                       if names_cards[key] in room6_doctors:
                          room6_doctors.remove(names_cards[key])
                    elif value == "nurse":
                       nurses.append(names_cards[key])
                       if names_cards[key] in room6_nurses:
                         room6_nurses.remove(names_cards[key])
                    elif value == "patient":
                       patients.append(names_cards[key])
                       if names_cards[key] in room6_patients:
                         room6_patients.remove(names_cards[key])
                    elif value == "device":  
                       devices.append(names_cards[key])
                       if names_cards[key] in room6_devices:
                         room6_devices.remove(names_cards[key])
                    else:
                       print(value)
                       print("error")
          
            # Μέγεθος τετραγώνου και θέση
          w  = details["size"][0] - 10
          h  = details["size"][1] - 10        
          x  = (details["size"][0] - w)//2 + details["position"][0] 
          y  = (details["size"][1] - h)//2 +details["position"][1] 
                 
          rectangle = pygame.Rect(x,y,w,h)
          pygame.draw.rect(screen, GREW, rectangle)
     
          font_title = pygame.font.SysFont(None, 30)
          font_in = pygame.font.SysFont(None, 25)

          text_surface1 = font_title.render('DOCTORS :', True, DOC)
          text_surface2 = font_title.render('NURSES :', True, NUR)
          text_surface3 = font_title.render('PATIENTS :', True, PAT)
          text_surface4 = font_title.render('DEVICES :', True, DEV)
          
          if room != "room6":
              doc = ' | '.join(map(str,doctors))
              nur = ' | '.join(map(str,nurses))
              pat = ' | '.join(map(str,patients))
              dev = ' | '.join(map(str,devices))
          else:
              doc = ' | '.join(map(str,room6_doctors))
              nur = ' | '.join(map(str,room6_nurses))
              pat = ' | '.join(map(str,room6_patients))
              dev = ' | '.join(map(str,room6_devices))

          text_surface11 = font_in.render(doc, True, DOCC)
          text_surface22 = font_in.render(nur, True, NURR)
          text_surface33 = font_in.render(pat, True, PATT)
          text_surface44 = font_in.render(dev, True, DEVV)

          if room!="room6":   
               text_rect1  = text_surface1.get_rect(topleft =(details["position"][0]+10,details["position"][1] + 20))
               text_rect11 = text_surface11.get_rect(topleft =(details["position"][0]+20,details["position"][1] + 40))
               text_rect2 =  text_surface2.get_rect(topleft =(details["position"][0]+10,details["position"][1] + 70))
               text_rect22 = text_surface22.get_rect(topleft =(details["position"][0]+20,details["position"][1] + 90))

               text_rect3 = text_surface3.get_rect(topleft =(details["position"][0]+10,details["position"][1] + 120))
               text_rect33 = text_surface33.get_rect(topleft =(details["position"][0]+20,details["position"][1] + 140))
               text_rect4 = text_surface4.get_rect(topleft =(details["position"][0]+10,details["position"][1] + 170))
               text_rect44 = text_surface44.get_rect(topleft =(details["position"][0]+20,details["position"][1] + 190))
          else:
               text_rect1  = text_surface1.get_rect(topleft =(details["position"][0]+10,details["position"][1] + 7))
               text_rect11 = text_surface11.get_rect(topleft =(details["position"][0]+20,details["position"][1] + 27))
               text_rect2 =  text_surface2.get_rect(topleft =(details["position"][0]+10,details["position"][1] + 52))
               text_rect22 = text_surface22.get_rect(topleft =(details["position"][0]+20,details["position"][1] + 72))

               text_rect3 = text_surface3.get_rect(topleft =(details["position"][0]+510,details["position"][1] + 7))
               text_rect33 = text_surface33.get_rect(topleft =(details["position"][0]+520,details["position"][1] + 27))
               text_rect4 = text_surface4.get_rect(topleft =(details["position"][0]+510,details["position"][1] + 52))
               text_rect44 = text_surface44.get_rect(topleft =(details["position"][0]+520,details["position"][1] + 72))

              
             
          screen.blit(text_surface3, text_rect3)	
          screen.blit(text_surface4, text_rect4)
          screen.blit(text_surface33, text_rect33)
          screen.blit(text_surface44, text_rect44)
          screen.blit(text_surface1, text_rect1)  # Εφαρμογή του πρώτου κειμένου
          screen.blit(text_surface2, text_rect2)
          screen.blit(text_surface11, text_rect11)  
          screen.blit(text_surface22, text_rect22)
        
def writeExcel(room,room_card_list,cards):

    
     set1 = set(room_card_list["room"+room])
     set2 = set(cards.split('|'))
     new_insert = list(set2 - set1)
     new_exit = list(set1 - set2)
     
     data=[]

     if (new_insert and new_insert != ['']) or (new_exit and new_exit != ['']):
          # Σύνδεση στη βάση δεδομένων
          try:
             mydb = mysql.connector.connect(
             host="localhost",
             user="root",
             password="root",
             database="diaxeirisi_poron" 
             )
             print("Επιτυχής σύνδεση στη βάση δεδομένων.")
             mycursor = mydb.cursor()
             
          except mysql.connector.Error as e:
             print("Σφάλμα κατά τη σύνδεση στη βάση δεδομένων:", e)

     
     if new_insert and new_insert != ['']:
          for el in new_insert:
              print("onma  poy mpike :  ",names_cards[el],"| idiotita  : ",findType(el))
              current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # Παίρνουμε την τρέχουσα ημερομηνία και ώρα
              
              if room !="6":
                   data.append([current_datetime,names_cards[el],findType(el),"6","EXIT"])
               
                   #Εισαγωγή δεδομένων στον πίνακα
                   sql = "INSERT INTO my_table(date_time, name, quality, room, action) VALUES (%s, %s, %s, %s, %s)"
                   val = (current_datetime, names_cards[el], findType(el),"6", "EXIT")
                   try:
                     mycursor.execute(sql, val)
                     mydb.commit()
                     
                   except mysql.connector.Error as e:
                       print("Σφάλμα κατά την εισαγωγή δεδομένων:", e)

              data.append([current_datetime,names_cards[el],findType(el),room,"INSERT"])
              #Εισαγωγή δεδομένων στον πίνακα
              sql = "INSERT INTO my_table(date_time, name, quality, room, action) VALUES (%s, %s, %s, %s, %s)"
              val = (current_datetime, names_cards[el], findType(el), room, "INSERT")
              try:
                mycursor.execute(sql, val)
                mydb.commit()
                
              except mysql.connector.Error as e:
                  print("Σφάλμα κατά την εισαγωγή δεδομένων:", e)

            
     
     if new_exit and new_exit!= ['']:
         for el in new_exit:
             print("onma  poy vgike :  ",names_cards[el],"| idiotita  : ",findType(el))
             current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")    # Παίρνουμε την τρέχουσα ημερομηνία και ώρα
             data.append([current_datetime,names_cards[el],findType(el),room,"EXIT"])

             #Εισαγωγή δεδομένων στον πίνακα
             sql = "INSERT INTO my_table(date_time, name, quality, room, action) VALUES (%s, %s, %s, %s, %s)"
             val = (current_datetime, names_cards[el], findType(el), room, "EXIT")
             try:
               mycursor.execute(sql, val)
               mydb.commit()
               
             except mysql.connector.Error as e:
                 print("Σφάλμα κατά την εισαγωγή δεδομένων:", e)

             if room !="6":
                   data.append([current_datetime,names_cards[el],findType(el),"6","INSERT"])

                   #Εισαγωγή δεδομένων στον πίνακα
                   sql = "INSERT INTO my_table(date_time, name, quality, room, action) VALUES (%s, %s, %s, %s, %s)"
                   val = (current_datetime, names_cards[el], findType(el),"6", "INSERT")
                   try:
                     mycursor.execute(sql, val)
                     mydb.commit()
                     
                   except mysql.connector.Error as e:
                       print("Σφάλμα κατά την εισαγωγή δεδομένων:", e)

             

     
     if data and data!= ['']:
          excel_file  = "log_temp.xlsx"
          try:
                wb = load_workbook(excel_file)
                ws = wb.active
                print("Το αρχείο Excel βρέθηκε και φορτώθηκε.")
          except FileNotFoundError:
                print("Το αρχείο Excel δεν βρέθηκε. Δημιουργία νέου αρχείου Excel...")
                wb = Workbook()
                ws = wb.active
                ws.append(["DATE-TIME", "NAME","QUALITY" , "ROOM", "ACTION"])
                print("Νέο αρχείο Excel δημιουργήθηκε.")
     
          for row in data:
              print(row)
              ws.append(row)
     
          wb.save(excel_file)
          wb.close()
          mydb.close()
     

def deleteDublicate(room,room_card_list,cards):

     global _id_for_delete
     set1 = set(room_card_list["room"+room])
     set2 = set(cards.split('|'))
     new_insert = list(set2 - set1)
     new_exit = list(set1 - set2)

     print("new_insert is : ",new_insert,"from room : ",room)
     if room == "1":
          if new_insert:
               for elem in new_insert:
                    if elem in room_card_list["room2"]:
                          #del room_card_list["room2"][elem]
                          _id_for_delete = "room2" + "-" + elem
                         
                          
                          #ser.write(new_insert.encode())
                    elif elem in room_card_list["room3"]:
                          #del room_card_list["room3"][elem]
                          _id_for_delete= "room3" + "-" + elem

    

     if room== "2":
          if new_insert:
               for elem in new_insert:
                    if elem in room_card_list["room1"]:
                          #del room_card_list["room1"][elem]
                          _id_for_delete= "room1" + "-" + elem

                    if elem in room_card_list["room3"]:
                          #del room_card_list["room2"][elem]
                          _id_for_delete= "room3" + "-" + elem

     if room== "3":
          if new_insert:
               for elem in new_insert:
                    if elem in room_card_list["room1"]:
                          #del room_card_list["room1"][elem]
                          _id_for_delete= "room1" + "-" + elem

                    if elem in room_card_list["room2"]:
                          #del room_card_list["room2"][elem]
                          _id_for_delete= "room2" + "-" + elem

     if room== "6":
    
          if new_exit:
              for elem in new_exit:
                  if elem in room_card_list["room1"]:
                          #del room_card_list["room1"][elem]
                          _id_for_delete= "room1" + "-" + elem

                  if elem in room_card_list["room2"]:
                          #del room_card_list["room2"][elem]
                          _id_for_delete= "room2" + "-" + elem

                  if elem in room_card_list["room3"]:
                          #del room_card_list["room3"][elem]
                          _id_for_delete= "room3" + "-" + elem

'''
     # Δημιουργία τετραγώνου
     pygame.draw.rect(screen, rectangle_color, (rectangle_x, rectangle_y, rectangle_width, rectangle_height),border_radius=10)

            # Κείμενο
     
     
     
     # Εφαρμογή κειμένου στο κέντρο της οθόνης
     

     screen.blit(text_surface1, text_rect1)  # Εφαρμογή του πρώτου κειμένου
     screen.blit(text_surface2, text_rect2)  # Εφαρμογή του δεύτερου κειμένου
'''